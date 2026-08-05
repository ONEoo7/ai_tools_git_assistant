"""Reading and writing rule tables as spreadsheets.

The file a team hands over is a spreadsheet, not JSON, and it was not written
for this program: the header may sit under a title row, the columns may be
called ``Rule ID`` or ``rule_id`` or ``RULEID``, and there may be four other
columns nobody here cares about. So the header is *found* rather than assumed,
and everything outside the two columns is ignored.

``openpyxl`` is imported inside the functions, as ``claude_client`` does with
``anthropic``: the tray menu must not pay to read a spreadsheet it will never
open. That is also why "openpyxl" is listed in every spec's ``hiddenimports`` --
PyInstaller cannot see an import that happens inside a function.
"""

from __future__ import annotations

from pathlib import Path

from git_assistant.review.rules import Rule, RuleTable, now_stamp

#: How far down a sheet the header row is looked for. Deeper than that and the
#: file is not a rules table with a title on top, it is something else.
HEADER_SEARCH_ROWS = 20

ID_HEADER = "ruleID"
DETAILS_HEADER = "ruleDetails"

#: Accepted spellings, normalized. The first pair is what is documented; the
#: rest are what real exports look like.
_ID_NAMES = ("ruleid", "id", "rule")
_DETAILS_NAMES = ("ruledetails", "details", "description", "ruletext", "rule")


class XlsxError(RuntimeError):
    """A spreadsheet that could not be read as rules, with the reason why."""


def _norm(value: object) -> str:
    """Header cells compared without case, spaces, dashes or underscores."""
    return "".join(c for c in str(value or "").lower() if c.isalnum())


def _cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _find_header(rows: list[tuple]) -> tuple[int, int, int]:
    """Locate ``(row index, id column, details column)`` in the first rows.

    Both columns must be found in the *same* row: a sheet where "details"
    happens to appear three rows below "ruleID" is not a table.
    """
    for index, row in enumerate(rows[:HEADER_SEARCH_ROWS]):
        headers = {_norm(cell): col for col, cell in enumerate(row) if _cell(cell)}
        id_col = next((headers[n] for n in _ID_NAMES if n in headers), None)
        details_col = next(
            (headers[n] for n in _DETAILS_NAMES if n in headers and headers[n] != id_col),
            None,
        )
        if id_col is not None and details_col is not None:
            return index, id_col, details_col
    raise XlsxError(
        f"no header row with a '{ID_HEADER}' column and a '{DETAILS_HEADER}' "
        f"column was found in the first {HEADER_SEARCH_ROWS} rows"
    )


def read_rules(path: str | Path, name: str = "") -> RuleTable:
    """Read one sheet of a workbook as a rule table.

    Every worksheet is tried, in order, and the first one carrying both columns
    wins -- a workbook whose first sheet is a cover page still imports. Rows
    without a rule id are skipped: those are the blank ones and the section
    headings people leave between groups of rules.
    """
    try:
        import openpyxl
    except ImportError as exc:  # a build that shipped without it
        raise XlsxError(f"reading spreadsheets needs openpyxl: {exc}") from exc

    path = Path(path)
    try:
        book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises a zoo of types for a bad file
        raise XlsxError(f"{path.name} could not be opened as a spreadsheet: {exc}") from exc

    try:
        problem = ""
        for sheet in book.worksheets:
            rows = [tuple(r) for r in sheet.iter_rows(values_only=True)]
            try:
                start, id_col, details_col = _find_header(rows)
            except XlsxError as exc:
                problem = problem or str(exc)
                continue
            rules: list[Rule] = []
            for row in rows[start + 1 :]:
                rule_id = _cell(row[id_col]) if id_col < len(row) else ""
                if not rule_id:
                    continue
                details = _cell(row[details_col]) if details_col < len(row) else ""
                rules.append(Rule(rule_id=rule_id, details=details))
            if not rules:
                problem = f"'{sheet.title}' has the columns but no rules under them"
                continue
            return RuleTable(
                name=(name or path.stem).strip() or path.stem,
                rules=rules,
                source=str(path),
                imported_at=now_stamp(),
            )
    finally:
        book.close()

    raise XlsxError(problem or f"{path.name} holds no readable rules")


def write_rules(path: str | Path, table: RuleTable) -> None:
    """Write a rule table as a spreadsheet this module can read back."""
    try:
        import openpyxl
    except ImportError as exc:
        raise XlsxError(f"writing spreadsheets needs openpyxl: {exc}") from exc

    book = openpyxl.Workbook()
    sheet = book.active
    # Sheet names are capped at 31 characters and cannot hold []:*?/\.
    sheet.title = "".join(c for c in table.name if c not in "[]:*?/\\")[:31] or "Rules"
    sheet.append([ID_HEADER, DETAILS_HEADER])
    for rule in table.rules:
        sheet.append([rule.rule_id, rule.details])
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 90
    try:
        book.save(str(path))
    except Exception as exc:
        raise XlsxError(f"could not write {path}: {exc}") from exc
    finally:
        book.close()
